"""
O.L.E.A. — Oltek Logistics Extraction Automation
backend/main.py

FastAPI application — two endpoints:
  POST /upload   → accepts PDF upload, returns parsed JSON with confidence data
  POST /export   → accepts parsed JSON, returns .xlsx download

Run locally:
  uvicorn main:app --reload --port 8000
"""

import io
import os
import tempfile
import time

from fastapi import FastAPI, File, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import BaseModel
from typing import Any

from parser import parse_contract


# ---------------------------------------------------------------------------
# App setup
# ---------------------------------------------------------------------------

app = FastAPI(
    title="O.L.E.A. API",
    description="Oltek Logistics Extraction Automation — zero-cost local PDF parsing",
    version="1.0.0",
)

# Allow the Vite dev server (port 5173) and any localhost origin
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://127.0.0.1:5173", "http://localhost:3000"],
    allow_methods=["*"],
    allow_headers=["*"],
)


# ---------------------------------------------------------------------------
# POST /upload — parse a PDF and return structured JSON
# ---------------------------------------------------------------------------

@app.post("/upload")
async def upload_pdf(file: UploadFile = File(...)):
    """
    Accept a PDF upload, run it through the extraction engine, return JSON.

    Response shape:
    {
        "processing_time_ms": 1243.5,
        "api_cost_usd": 0.00,
        "overall_confidence": 0.94,
        "page_count": 94,
        "filename": "ATL0347N25.pdf",
        "header": { contract_no, carrier, shipper, effective_date, expiration_date, ... },
        "commodities": [ { description: {...} }, ... ],
        "rate_records": [ { origin, destination, terminal, rate_20, ... }, ... ],
    }
    """
    # Validate file type
    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="Only PDF files are accepted.")

    # Write upload to a temp file (pdfplumber needs a real file path)
    content = await file.read()
    if len(content) == 0:
        raise HTTPException(status_code=400, detail="Uploaded file is empty.")

    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
        tmp.write(content)
        tmp_path = tmp.name

    try:
        result = parse_contract(tmp_path)
        result["filename"] = file.filename
        return result
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Parsing failed: {str(exc)}")
    finally:
        # Always clean up the temp file
        os.unlink(tmp_path)


# ---------------------------------------------------------------------------
# POST /export — convert parsed data to Excel (.xlsx) download
# ---------------------------------------------------------------------------

class ExportPayload(BaseModel):
    """The full JSON object returned from /upload — sent back for Excel export."""
    filename: str = "contract"
    header: dict[str, Any]
    commodities: list[dict[str, Any]]
    rate_records: list[dict[str, Any]]
    processing_time_ms: float = 0
    api_cost_usd: float = 0.00
    overall_confidence: float = 0
    page_count: int = 0


def _unwrap(v: Any) -> Any:
    """Extract the raw .value from a confidence-wrapped field dict, else passthrough."""
    if isinstance(v, dict) and "value" in v:
        return v["value"]
    return v


@app.post("/export")
async def export_excel(payload: ExportPayload):
    """
    Build an Excel workbook that exactly matches the OLTK service contract
    template (ATL0347N25 Template).  Single sheet 'Rates' with the same
    20 columns as the reference template.
    """
    # Pull header values once
    carrier     = _unwrap(payload.header.get("carrier",        {})) or "OLTEK"
    contract_id = _unwrap(payload.header.get("contract_no",    {})) or ""
    eff_date    = _unwrap(payload.header.get("effective_date",  {})) or ""
    exp_date    = _unwrap(payload.header.get("expiration_date", {})) or ""

    wb = Workbook()
    ws = wb.active
    ws.title = "Rates"

    # ------------------------------------------------------------------
    # Header row — exact column order matching the template
    # ------------------------------------------------------------------
    HEADERS = [
        "Carrier",              # A
        "Contract ID",          # B
        "effective_date",       # C
        "expiration_date",      # D
        "commodity",            # E
        "origin_city",          # F
        "origin_via_city",      # G  (not extracted — blank)
        "destination_city",     # H
        "destination_via_city", # I
        "service",              # J
        "Remarks",              # K  (blank)
        "SCOPE",                # L
        "BaseRate 20",          # M
        "BaseRate 40",          # N
        "BaseRate 40H",         # O
        "BaseRate 45",          # P
    ]

    # Column widths (characters)
    COL_WIDTHS = [8, 13, 13, 15, 32, 20, 20, 20, 20, 10, 12, 32, 12, 12, 12, 12]

    # Header cell style
    hdr_fill  = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    hdr_font  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_font = Font(name="Calibri", size=10)

    ws.append(HEADERS)
    ws.row_dimensions[1].height = 32
    for col_idx, cell in enumerate(ws[1], start=1):
        cell.fill  = hdr_fill
        cell.font  = hdr_font
        cell.alignment = hdr_align
        ws.column_dimensions[cell.column_letter].width = COL_WIDTHS[col_idx - 1]

    # ------------------------------------------------------------------
    # Data rows
    # ------------------------------------------------------------------
    for rec in payload.rate_records:
        terminal = _unwrap(rec.get("terminal", {})) or ""
        service  = f"{terminal}/CY" if terminal else ""

        row = [
            carrier,
            contract_id,
            eff_date,
            exp_date,
            _unwrap(rec.get("commodity",    {})),   # E
            _unwrap(rec.get("origin",       {})),   # F  origin_city
            None,                                   # G  origin_via_city — not extracted
            _unwrap(rec.get("destination",  {})),   # H  destination_city
            _unwrap(rec.get("via",          {})),   # I  destination_via_city
            service,                                # J
            None,                                   # K  Remarks
            _unwrap(rec.get("trade_lane",   {})),   # L  SCOPE
            _unwrap(rec.get("rate_20",      {})),   # M
            _unwrap(rec.get("rate_40",      {})),   # N
            _unwrap(rec.get("rate_40hc",    {})),   # O
            _unwrap(rec.get("rate_45",      {})),   # P
        ]
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.font = data_font

    # ------------------------------------------------------------------
    # Freeze the header row, auto-filter on row 1
    # ------------------------------------------------------------------
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:P1"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    safe_name     = os.path.splitext(payload.filename)[0] or "contract"
    download_name = f"{safe_name}_extracted.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{download_name}"'},
    )


# ---------------------------------------------------------------------------
# Health-check
# ---------------------------------------------------------------------------

@app.get("/health")
def health():
    return {"status": "ok", "service": "O.L.E.A.", "api_cost_usd": 0.00}
